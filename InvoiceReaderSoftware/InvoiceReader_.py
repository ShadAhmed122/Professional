import PyPDF2 as p
from datetime import datetime
import openpyxl
import re,os

def extract_invoice_number(text):
    match = re.search(r"Invoice No: ([A-Z]+-\d[\d\s]*)", text)
    if match:
        return match.group(1).replace(" ", "")
    return "N/A"

def extract_date(text):
    match = re.search(r"\b\d{2}-\d{2}-\d{4}\b", text)
    if match:
        return match.group(0)
    return "N/A"

def extract_shipping_cost(text):
    match = re.search(r"Shipping Cost Tk ([\d,]+\.\d+)", text)
    if match:
        return int(float(match.group(1).replace(',', '')))
    return 0

def extract_product_info(text):

    product_pattern = re.compile(r"([A-Za-z ]+ [\d.]+ [A-Za-z]+) (\d+) Tk ([\d,]+\.\d{2}) Tk ([\d,]+\.\d{2})")
    products = product_pattern.findall(text)
    product_details = []
    
    for product in products:
        product_name = product[0].strip()
        qty = int(product[1].strip())
        price = int(float(product[2].replace(',', '')))
        total = int(float(product[3].replace(',', '')))
        
   
        product_details.append({
            'Product Name': product_name,
            'Quantity': qty,
            'Price': price,
            'Total': total
        })
    
 
    return product_details, len(products)

def extract_discount(text):
    match = re.search(r"Discount Tk ([\d,]+\.\d+)", text)
    if match:
        return int(float(match.group(1).replace(',', '')))
    return 0

# def extract_discount(text):
#     discount_pattern = r"Discount Tk (\d+(?:\.\d{1,2})?)"
#     discount_match = re.search(discount_pattern, text)
#     if discount_match:
#         discount = int(float(discount_match.group(1)))
#     else:
#         discount = 0
#     return discount


if __name__ == '__main__':
    current_dir = os.getcwd()
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    os.chdir(desktop_path)
    
    
    pdf_file = None
    for file in os.listdir():
        if file.endswith(".pdf"):
            pdf_file = file
            break
        
    pdf_path = os.path.join(desktop_path, pdf_file)
    pdf_file_obj = open(pdf_path, "rb")
    pdf = p.PdfReader(pdf_file_obj)
    # pdf = p.PdfReader(open(r"D:\OfficeDocumentsOfShadAhmed\MyProfessionalRepository\Professional\InvoiceReaderSoftware\abc2.pdf", "rb"))
    pdf=pdf.pages[:]
    
    
    current_date = datetime.now()
    date_string = current_date.strftime("%d-%m-%Y")
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ConvertedOn_" +datetime.now().strftime("%d-%m-%Y")
    data = ['Invoice No.', 'Item Name','Sales Description','Selling Price','Quantity',
            'Discount', 'Shipping Charge', 'Final Invoice Price','Date', 'Product Type',
            'Sales Account','Order No.']
    
    # for row_num, row_data in enumerate(data, start=1): 
    row=1
    for col_num, value in enumerate(data, start=1):
        sheet.cell(row=row, column=col_num).value = value
    
    
    row=2
    items=0
    for i in pdf:
        a=i.extract_text()
        
        print(extract_invoice_number(a))
        sheet.cell(row=row, column=1).value = extract_invoice_number(a)
        
        print(extract_date(a))
        sheet.cell(row=row, column=9).value =extract_date(a)
        
        print(extract_shipping_cost(a))
        sheet.cell(row=row, column=7).value = extract_shipping_cost(a)
        
        print("Discount: ",extract_discount(a))
        sheet.cell(row=row, column=6).value = extract_discount(a)
        
        product_info, total_products = extract_product_info(a)
        # print(total_products)
        
        for k,product in enumerate(product_info):
            print(product['Product Name'])
            sheet.cell(row=row, column=2).value = product['Product Name']
            
            print(product['Quantity'])
            sheet.cell(row=row, column=5).value = product['Quantity']
            
            print(product['Price'])
            sheet.cell(row=row, column=4).value = product['Price']
            
            if k==0:
                print(product['Total']+extract_shipping_cost(a)-extract_discount(a))
                sheet.cell(row=row, column=8).value = product['Total']+extract_shipping_cost(a)-extract_discount(a)
            else:
                print(product['Total'])
                sheet.cell(row=row, column=8).value = product['Total']
                
            
            items +=1
            print(items)
            
            row+=1
            
            
        print(" ")
    workbook.save('new_excel_file2.xlsx')
    print("Excel file created successfully!")
    os.chdir(current_dir)
