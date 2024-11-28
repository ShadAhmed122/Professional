import PyPDF2 as p
import datetime 
import openpyxl
import re,os
import pygame,pyttsx3
from pygame.locals import *
import sys
import time

global current_dir



def excelll():

    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    os.chdir(desktop_path)
    
    
    pdf_file = None
    for file in os.listdir():
        if file.endswith(".pdf"):
            pdf_file = file
            break
    if pdf_file:
        pdf_file_name = os.path.splitext(pdf_file)[0]
    else:
        spek("No PDF file in the Desktop")
        return 1
    os.chdir(current_dir)
    print("pdf readed")
    pdf_path = os.path.join(desktop_path, pdf_file)
    pdf_file_obj = open(pdf_path, "rb")
    pdf = p.PdfReader(pdf_file_obj)
    # pdf = p.PdfReader(open(r"D:\OfficeDocumentsOfShadAhmed\MyProfessionalRepository\Professional\InvoiceReaderSoftware\abc2.pdf", "rb"))
    pdf=pdf.pages[:]
    
    
    # current_date =datetime.datetime.now()
    # date_string = current_date.strftime("%d-%m-%Y")
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ConvertedOn_" +datetime.datetime.now().strftime("%d-%m-%Y")
    data = ['Invoice No.', 'Item Name','Sales Description','Selling Price','Quantity',
            'Discount', 'Shipping Charge', 'Final Invoice Price','Date', 'Product Type',
            'Sales Account','Order No.']
    
    # for row_num, row_data in enumerate(data, start=1): 
    row=1
    for col_num, value in enumerate(data, start=1):
        sheet.cell(row=row, column=col_num).value = value
    
    print("work bench ")
    row=2
    items=0
    for i in pdf:
        a=i.extract_text()
        
        print(extract_invoice_number(a))
        sheet.cell(row=row, column=1).value = extract_invoice_number(a)
        
        print(extract_date(a))
        sheet.cell(row=row, column=9).value =extract_date(a)
        
        print(extract_shipping_cost(a))
        sheet.cell(row=row, column=7).value = extract_shipping_cost(a)
        
        print("Discount: ",extract_discount(a))
        sheet.cell(row=row, column=6).value = extract_discount(a)
        
        product_info, total_products = extract_product_info(a)
        # print(total_products)
        
        for k,product in enumerate(product_info):
            print(product['Product Name'])
            sheet.cell(row=row, column=2).value = product['Product Name']
            
            print(product['Quantity'])
            sheet.cell(row=row, column=5).value = product['Quantity']
            
            print(product['Price'])
            sheet.cell(row=row, column=4).value = product['Price']
            
            if k==0:
                print(product['Total']+extract_shipping_cost(a)-extract_discount(a))
                sheet.cell(row=row, column=8).value = product['Total']+extract_shipping_cost(a)-extract_discount(a)
            else:
                print(product['Total'])
                sheet.cell(row=row, column=8).value = product['Total']
                
            
            items +=1
            print(items)
            
            row+=1
            
            
        print(" ")
        
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    os.chdir(desktop_path)
    # workbook.save('new_excel_file2.xlsx')
    
    # excel_file_name = pdf_file_name + '.xlsx'
    excel_file_name = "Extracted_Data" + '.xlsx'
    workbook.save(excel_file_name)
    os.chdir(current_dir)
    print("Excel file created successfully!")
   
    
    return 0


def extract_invoice_number(text):
    match = re.search(r"Invoice No: ([A-Z]+-\d[\d\s]*)", text)
    if match:
        return match.group(1).replace(" ", "")
    return "N/A"

def extract_date(text):
    match = re.search(r"\b\d{2}-\d{2}-\d{4}\b", text)
    if match:
        return match.group(0)
    return "N/A"

def extract_shipping_cost(text):
    match = re.search(r"Shipping Cost Tk ([\d,]+\.\d+)", text)
    if match:
        return int(float(match.group(1).replace(',', '')))
    return 0

def extract_product_info(text):

    product_pattern = re.compile(r"([A-Za-z ]+ [\d.]+ [A-Za-z]+) (\d+) Tk ([\d,]+\.\d{2}) Tk ([\d,]+\.\d{2})")
    products = product_pattern.findall(text)
    product_details = []
    
    for product in products:
        product_name = product[0].strip()
        qty = int(product[1].strip())
        price = int(float(product[2].replace(',', '')))
        total = int(float(product[3].replace(',', '')))
        
   
        product_details.append({
            'Product Name': product_name,
            'Quantity': qty,
            'Price': price,
            'Total': total
        })
    
 
    return product_details, len(products)

def extract_discount(text):
    match = re.search(r"Discount Tk ([\d,]+\.\d+)", text)
    if match:
        return int(float(match.group(1).replace(',', '')))
    return 0


def spek(audio):
    speak=pyttsx3.init("sapi5")
    voice= speak.getProperty('voices')
    speak.setProperty('voice', voice[1].id)
    speak.say(audio)
    speak.runAndWait()
    print(audio)  
    
def wishMe():
    hour = int(datetime.datetime.now().hour)
    if hour>=5 and hour<12:
        spek("...........................................  Good Morning! ")
    elif hour>=12 and hour<16:
        spek("...........................................  Good Afternoon!")   
    elif hour>=16 and hour<19:
        spek(".............................................  Good Evening! ") 
    else:
        spek('..............................................  Good night')

def Welcome():
    welcome_image = pygame.image.load('welcome.png').convert_alpha()
    welcome_image = pygame.transform.scale(welcome_image, (1535, 865))
    screen.blit(welcome_image, (193,110))
    pygame.display.update()
    wishMe()

    while True:
        for event in pygame.event.get():
            if event.type == QUIT or (event.type == KEYDOWN and event.key == K_ESCAPE):
                pygame.quit()
                sys.exit()
            elif event.type == MOUSEBUTTONDOWN or event.type == KEYDOWN:
                return
            else:
                screen.blit(welcome_image, (193,110))
                pygame.display.update()
                fps.tick(60)   
                           
def Guideline():
    welcome_image = pygame.image.load('guideline.png').convert_alpha()
    welcome_image = pygame.transform.scale(welcome_image, (1535, 865))
    screen.blit(welcome_image, (193,110))
    pygame.display.update()
    
    spek("Extracting data from a pdf is very easy with this software. All you have to do is just copy the target pdf file into the Desktop. Avoid keeping any other PDF file on the desktop. Now just hit the extract button and done. You will get you’re extracted data converted into an Excel file.")

    while True:
        for event in pygame.event.get():
            if event.type == QUIT or (event.type == KEYDOWN and event.key == K_ESCAPE):
                pygame.quit()
                sys.exit()
            screen.blit(welcome_image, (193,110))
            mouse = pygame.mouse.get_pos()
            
            
            color = (255, 255, 255)
            color_light = (0, 0, 0)
            color_dark = (100, 100, 100)
            smallfont = pygame.font.SysFont('Corbel', 75)
            text = smallfont.render(' Back', True, color)
            
            
            
            x = 794
            y = 794
            width = 346
            height = 119
            if event.type == pygame.MOUSEBUTTONDOWN:
                if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                    mainloop()

            if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                pygame.draw.rect(screen, color_light, [x, y, width, height])
                text = smallfont.render(' Back', True, (160, 32, 240))
                screen.blit(text, (x+80, y+30))
            else:
                pygame.draw.rect(screen, color_dark, [x, y, width, height])
                screen.blit(text, (x+80, y+30))
            
            
            
            
            
            pygame.display.update()
            fps.tick(60)                
      
def Exxx():
    welcome_image = pygame.image.load('extracting.png').convert_alpha()
    welcome_image = pygame.transform.scale(welcome_image, (1535, 865))
    screen.blit(welcome_image, (193,110))
    pygame.display.update()
    # time.sleep(3)
    spek("Extracting Data from PDF file")
    
    try:
        a=excelll()
        if a==1:
            welcome_image = pygame.image.load('fail.png').convert_alpha()
            welcome_image = pygame.transform.scale(welcome_image, (1535, 865))
            while True:
                for event in pygame.event.get():
                    if event.type == QUIT or (event.type == KEYDOWN and event.key == K_ESCAPE):
                        pygame.quit()
                        sys.exit()
                    screen.blit(welcome_image, (193,110))
                    mouse = pygame.mouse.get_pos()
                    
                    
                    color = (255, 255, 255)
                    color_light = (0, 0, 0)
                    color_dark = (100, 100, 100)
                    smallfont = pygame.font.SysFont('Corbel', 75)
                    
                    
                    
                    
                    
                    x = 537
                    y = 679
                    width = 281
                    height = 113
                    text = smallfont.render('Back', True, color)
                    if event.type == pygame.MOUSEBUTTONDOWN:
                        if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                            mainloop()

                    if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                        pygame.draw.rect(screen, color_light, [x, y, width, height])
                        text = smallfont.render('Back', True, (160, 32, 240))
                        screen.blit(text, (x+70, y+30))
                    else:
                        pygame.draw.rect(screen, color_dark, [x, y, width, height])
                        screen.blit(text, (x+70, y+30))
                        
                        
                        
                        
                    x = 1039
                    y = 679
                    width = 294
                    height = 103
                    text = smallfont.render('  Exit', True, color)
                    if event.type == pygame.MOUSEBUTTONDOWN:
                        if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                            pygame.quit()
                            sys.exit()

                    if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                        pygame.draw.rect(screen, color_light, [x, y, width, height])
                        text = smallfont.render('  Exit', True, (160, 32, 240))
                        screen.blit(text, (x+60, y+20))
                    else:
                        pygame.draw.rect(screen, color_dark, [x, y, width, height])
                        screen.blit(text, (x+60, y+20))
                    
                    
                    
                    
                    
                    
                    
                    pygame.display.update()
                    fps.tick(60)                
    

                
            
        welcome_image = pygame.image.load('congo.png').convert_alpha()
        welcome_image = pygame.transform.scale(welcome_image, (1535, 865))
        screen.blit(welcome_image, (193,110))
        pygame.display.update()
        spek("Congratulations. Your Excel File is created")
        while True:
        
            for event in pygame.event.get():
                if event.type == QUIT or (event.type == KEYDOWN and event.key == K_ESCAPE):
                    pygame.quit()
                    sys.exit()
                screen.blit(welcome_image, (193,110))
                mouse = pygame.mouse.get_pos()
                
                
                color = (255, 255, 255)
                color_light = (0, 0, 0)
                color_dark = (100, 100, 100)
                smallfont = pygame.font.SysFont('Corbel', 75)
                
                
                
                
                x = 497
                y = 639
                width = 303
                height = 124
                text = smallfont.render('Back', True, color)
                if event.type == pygame.MOUSEBUTTONDOWN:
                    if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                        mainloop()

                if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                    pygame.draw.rect(screen, color_light, [x, y, width, height])
                    text = smallfont.render('Back', True, (160, 32, 240))
                    screen.blit(text, (x+80, y+30))
                else:
                    pygame.draw.rect(screen, color_dark, [x, y, width, height])
                    screen.blit(text, (x+80, y+30))
                    
                    
                    
                    
                x = 1184
                y = 639
                width = 285
                height = 127
                text = smallfont.render('Exit', True, color)
                if event.type == pygame.MOUSEBUTTONDOWN:
                    if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                        pygame.quit()
                        sys.exit()

                if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                    pygame.draw.rect(screen, color_light, [x, y, width, height])
                    text = smallfont.render('Exit', True, (160, 32, 240))
                    screen.blit(text, (x+80, y+30))
                else:
                    pygame.draw.rect(screen, color_dark, [x, y, width, height])
                    screen.blit(text, (x+80, y+30))
                
                
                
                
                
                pygame.display.update()
                fps.tick(60)                
    except Exception as e:
        print(e)
        os.chdir(current_dir)
        welcome_image = pygame.image.load('fail.png').convert_alpha()
        welcome_image = pygame.transform.scale(welcome_image, (1535, 865))
        while True:
            for event in pygame.event.get():
                if event.type == QUIT or (event.type == KEYDOWN and event.key == K_ESCAPE):
                    pygame.quit()
                    sys.exit()
                screen.blit(welcome_image, (193,110))
                mouse = pygame.mouse.get_pos()
                
                
                color = (255, 255, 255)
                color_light = (0, 0, 0)
                color_dark = (100, 100, 100)
                smallfont = pygame.font.SysFont('Corbel', 75)
                
                
                
                
                
                x = 537
                y = 679
                width = 281
                height = 113
                text = smallfont.render('Back', True, color)
                if event.type == pygame.MOUSEBUTTONDOWN:
                    if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                        mainloop()

                if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                    pygame.draw.rect(screen, color_light, [x, y, width, height])
                    text = smallfont.render('Back', True, (160, 32, 240))
                    screen.blit(text, (x+70, y+30))
                else:
                    pygame.draw.rect(screen, color_dark, [x, y, width, height])
                    screen.blit(text, (x+70, y+30))
                    
                    
                    
                    
                x = 1039
                y = 679
                width = 294
                height = 103
                text = smallfont.render('  Exit', True, color)
                if event.type == pygame.MOUSEBUTTONDOWN:
                    if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                        pygame.quit()
                        sys.exit()

                if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                    pygame.draw.rect(screen, color_light, [x, y, width, height])
                    text = smallfont.render('  Exit', True, (160, 32, 240))
                    screen.blit(text, (x+60, y+20))
                else:
                    pygame.draw.rect(screen, color_dark, [x, y, width, height])
                    screen.blit(text, (x+60, y+20))
                
                
                
                
                
                
                
                pygame.display.update()
                fps.tick(60)                
 
def mainloop():
  
    homepage_image = pygame.image.load('Homepage.png').convert_alpha()
    homepage_image = pygame.transform.scale(homepage_image, (1535, 865)) 

    while True:
        for event in pygame.event.get():
            if event.type == pygame.QUIT or (event.type == pygame.KEYDOWN and event.key == pygame.K_ESCAPE):
                pygame.quit()
                sys.exit()

            # Display the homepage image
            screen.blit(homepage_image,(193,110))

            # Get mouse position
            mouse = pygame.mouse.get_pos()
            

            # Button 1: Start (right)
            color = (2, 48, 32)
            color_light = (250, 160, 160)
            color_dark = (100, 100, 100)
            smallfont = pygame.font.SysFont('Corbel', 55)
            text = smallfont.render('         Exit', True, color)
            x = 1273
            y = 601
            width = 283
            height = 119

            if event.type == pygame.MOUSEBUTTONDOWN:
                if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                    pygame.quit()
                    sys.exit()

            if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                pygame.draw.rect(screen, color_light, [x, y, width, height])
                screen.blit(text, (x, y+45))

            # Button 2: Start (left)
            text = smallfont.render(' Guideline', True, color)
            x = 328
            y = 608
            width = 265
            height = 114

            if event.type == pygame.MOUSEBUTTONDOWN:
                if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                    Guideline()

            if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                pygame.draw.rect(screen, color_light, [x, y, width, height])
                screen.blit(text, (x+15, y+40))

            # Button 3: Start (middle)
            text = smallfont.render('    Extract', True, color)
            x = 792
            y = 604
            width = 275
            height = 116

            if event.type == pygame.MOUSEBUTTONDOWN:
                if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                    Exxx()

            if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
                pygame.draw.rect(screen, color_light, [x, y, width, height])
                screen.blit(text, (x+15, y+40))

            # Button 4: Quit
            # color = (255, 255, 255)
            # color_light = (0, 0, 0)
            # color_dark = (100, 100, 100)
            # smallfont = pygame.font.SysFont('Corbel', 75)
            # text = smallfont.render(' Quit', True, color)
            # x = 880
            # y = 900
            # width = 150
            # height = 70

            # if event.type == pygame.MOUSEBUTTONDOWN:
            #     if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
            #         pygame.quit()

            # if x <= mouse[0] <= x+width and y <= mouse[1] <= y+height:
            #     pygame.draw.rect(screen, color_light, [x, y, width, height])
            #     text = smallfont.render(' Quit', True, (160, 32, 240))
            #     screen.blit(text, (x, y))
            # else:
            #     pygame.draw.rect(screen, color_dark, [x, y, width, height])
            #     screen.blit(text, (x, y))

            pygame.display.update()
            fps.tick(60)
      
if __name__=="__main__":
    current_dir = os.getcwd()
    screen= pygame.display.set_mode((1920,1080))
    pygame.init()
    fps=pygame.time.Clock()
    pygame.display.set_caption("WheelChair Controller")
    
    # while True:
    Welcome()
    mainloop()